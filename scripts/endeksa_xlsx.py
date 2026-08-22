"""Build a district workbook from an Endeksa dump plus TUIK neighbourhood series.

Usage:
    python scripts/endeksa_xlsx.py --district TR-16-006 [--dump endeksa-16-1420.json]
                                   [--raw C:/veri/raw/endeksa] [--out out.xlsx]

Layout: one topic per sheet, a two-row header (group / column), counts from the
source and every derived figure (share, density, mean age, turnout) as a formula
on the same row. Raw JSON stays the source of truth; see docs/endeksa.md.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as L
from openpyxl.worksheet.worksheet import Worksheet

REPO = Path(__file__).resolve().parent.parent
FONT = "Arial"
NAVY = "1F3864"
BLUE = "2F5597"
BAND = "F3F6FB"
SUBTOTAL = "DDE6F3"
GREY = "6B6B6B"
THIN = Side(style="thin", color="D0D7E2")
FMT_INT = "#,##0"
FMT_PCT = "0.0%"
FMT_DEC1 = "0.0"
FMT_DEC2 = "0.00"

AGE_BANDS = [
    "0_4",
    "5_9",
    "10_14",
    "15_19",
    "20_24",
    "25_29",
    "30_34",
    "35_39",
    "40_44",
    "45_49",
    "50_54",
    "55_59",
    "60_64",
    "65",
]
AGE_LABEL = {b: (b.replace("_", "-") if b != "65" else "65+") for b in AGE_BANDS}
# band midpoints for the mean-age estimate; open 65+ band set to 74 (TUIK 65+ mean age, 2024)
AGE_MID = [2, 7, 12, 17, 22, 27, 32, 37, 42, 47, 52, 57, 62, 74]
SEX = [("Total", "Toplam"), ("Male", "Erkek"), ("Female", "Kadın")]
EDU = [
    ("EduNonLiterated", "Okuma yazma bilmeyen"),
    ("EduLiteratedUntutored", "Bilen, okul bitirmemiş"),
    ("EduPrimarySchool", "İlkokul"),
    ("EduPrimaryEducation", "İlköğretim"),
    ("EduMiddleSchool", "Ortaokul"),
    ("EduHighSchool", "Lise"),
    ("EduLicenseDegree", "Lisans"),
    ("EduGraduate", "Yüksek lisans"),
    ("EduDoctorate", "Doktora"),
    ("EduUnknown", "Bilinmeyen"),
]
MARITAL = [
    ("MarriedNever", "Hiç evlenmedi"),
    ("Married", "Evli"),
    ("Divorced", "Boşanmış"),
    ("Widow", "Eşi ölmüş"),
]
SES = [
    ("SesGroupAPlus", "A+"),
    ("SesGroupA", "A"),
    ("SesGroupB", "B"),
    ("SesGroupC", "C"),
    ("SesGroupD", "D"),
]
EXPENSE = [
    ("ExpenseFood", "Gıda"),
    ("ExpenseAlcoholAndSmoking", "Alkol-tütün"),
    ("ExpenseClothing", "Giyim"),
    ("ExpenseShelter", "Barınma"),
    ("ExpenseFurniture", "Mobilya"),
    ("ExpenseHealth", "Sağlık"),
    ("ExpenseTransportation", "Ulaşım"),
    ("ExpenseCommunication", "İletişim"),
    ("ExpenseEntertainment", "Eğlence"),
    ("ExpenseEducation", "Eğitim"),
    ("ExpenseRestaurant", "Restoran"),
    ("ExpenseOther", "Diğer"),
]
ELECTION_LABEL = {
    "2011genelsecim": "2011 Genel",
    "2014cumhurbaskani": "2014 CB",
    "2014yerel": "2014 Yerel",
    "2015haziran": "2015 Haz. Genel",
    "2015kasim": "2015 Kas. Genel",
    "2017anayasa": "2017 Referandum",
    "2018cumhurbaskani": "2018 CB",
    "2018genel": "2018 Genel",
    "2019yerelseçimilçebelediye": "2019 İlçe Bld.",
    "2019yerelseçimbelediyemeclisi": "2019 Bld. Meclisi",
    "2019yerelseçimbüyükşehir": "2019 Büyükşehir",
    "2023genel": "2023 Genel",
    "2023CumhurTur1": "2023 CB 1. tur",
    "2023CumhurTur2": "2023 CB 2. tur",
    "2024yerelseçimbelediyebaşkanlığı": "2024 İlçe Bld.",
    "2024yerelseçimbelediyemeclisüyeliği": "2024 Bld. Meclisi",
    "2024yerelseçimbüyükşehirbelediyebaşkanlığı": "2024 Büyükşehir",
}
FOCUS_ELECTION = "2024yerelseçimbelediyebaşkanlığı"
KIND_LABEL = {"centre": "Merkez", "rural": "Kır"}


def slug(s: str) -> str:
    tr = str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosucgiosu")
    return re.sub(r"[^a-z0-9]+", "-", s.translate(tr).lower()).strip("-")


def norm_name(s: str) -> str:
    s = s.replace("I", "ı").replace("İ", "i").lower()
    return re.sub(r"\s*(mah\.|mahallesi|mah)$", "", s).strip()


# ---------------------------------------------------------------- loading


def unpack_dump(dump: Path, raw_dir: Path) -> None:
    d = json.loads(dump.read_text(encoding="utf-8"))
    raw_dir.mkdir(parents=True, exist_ok=True)
    meta = d["meta"]
    (raw_dir / "county.json").write_text(
        json.dumps({"_meta": meta, **d["county"]}, ensure_ascii=False), "utf-8"
    )
    for did, q in d["quarters"].items():
        name = q["Demography"]["DistrictName"]
        q["_meta"] = {
            "fetched": meta["fetched"],
            "placeholder": q["Demography"]["HouseholdCount"] == 0,
        }
        (raw_dir / f"{did}-{slug(name)}.json").write_text(
            json.dumps(q, ensure_ascii=False), "utf-8"
        )
    (raw_dir / "election.json").write_text(
        json.dumps({"_meta": meta, **d["election"]}, ensure_ascii=False), "utf-8"
    )
    (raw_dir / "fellowcountryman.json").write_text(
        json.dumps({"_meta": meta, **d["fellows"]}, ensure_ascii=False), "utf-8"
    )
    if d.get("geo"):
        (raw_dir / "geo.json").write_text(
            json.dumps({"_meta": meta, **d["geo"]}, ensure_ascii=False), "utf-8"
        )


def load_raw(raw_dir: Path) -> dict:
    county = json.loads((raw_dir / "county.json").read_text("utf-8"))
    quarters = {}
    for p in raw_dir.glob("*-*.json"):
        if p.name[0].isdigit():
            q = json.loads(p.read_text("utf-8"))
            quarters[str(q["Demography"]["DistrictId"])] = q["Demography"]
    election = json.loads((raw_dir / "election.json").read_text("utf-8"))
    fellows = json.loads((raw_dir / "fellowcountryman.json").read_text("utf-8"))
    geo_ids = set()
    gp = raw_dir / "geo.json"
    if gp.exists():
        geo_ids = {f["id"] for f in json.loads(gp.read_text("utf-8"))["features"]}
    return {
        "county": county,
        "quarters": quarters,
        "election": election,
        "fellows": fellows,
        "geo_ids": geo_ids,
    }


def load_tuik(district: str) -> tuple[dict[str, dict], dict[str, dict[int, dict]]]:
    areas = {}
    with open(
        REPO / "src/veriatlas/data/areas_tr_neighbourhoods.csv", encoding="utf-8"
    ) as f:
        for r in csv.DictReader(f):
            if r["parent_id"] == district:
                areas[norm_name(r["name_tr"])] = r
    series: dict[str, dict[int, dict]] = {}
    with gzip.open(
        REPO / "public/population-neighbourhood.csv.gz", "rt", encoding="utf-8"
    ) as f:
        for r in csv.DictReader(f):
            if not r["area_id"].startswith(district + "-"):
                continue
            cell = series.setdefault(r["area_id"], {}).setdefault(int(r["year"]), {})
            cell[(r["age"] or "all") + "|" + (r["sex"] or "all")] = int(r["value"])
    return areas, series


def tuik_get(t: dict, y: int, age: str = "all", sex: str = "all"):
    cell = t.get(y, {})
    if age == "all" and sex == "all":
        return cell.get("all|all") or (
            sum(
                v
                for k, v in cell.items()
                if k.endswith("|all") and not k.startswith("all")
            )
            or None
        )
    if age == "all":
        return cell.get(f"all|{sex}") or (
            (cell.get(f"18+|{sex}", 0) + cell.get(f"0-17|{sex}", 0)) or None
        )
    return cell.get(f"{age}|{sex}")


def median_age(counts: list[int]) -> float | None:
    total = sum(counts)
    if total == 0:
        return None
    half = total / 2
    acc = 0
    edges = [0, 5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 85]
    for k, c in enumerate(counts):
        if acc + c >= half:
            lo, hi = edges[k], edges[k + 1]
            return round(lo + (half - acc) / c * (hi - lo), 1) if c else float(lo)
        acc += c
    return None


# ---------------------------------------------------------------- table writer


@dataclass
class Col:
    label: str
    fmt: str = FMT_INT
    width: float = 10
    scale: bool = False
    group: str = ""


@dataclass
class Table:
    ws: Worksheet
    title: str
    subtitle: str
    cols: list[Col]
    first_col: Col = field(default_factory=lambda: Col("Mahalle", "@", 24))
    header_row: int = 4
    rows_written: int = 0

    def __post_init__(self) -> None:
        ws = self.ws
        ws["A1"] = self.title
        ws["A1"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
        ws["A2"] = self.subtitle
        ws["A2"].font = Font(name=FONT, italic=True, size=9, color=GREY)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        g_row, h_row = self.header_row - 1, self.header_row
        allcols = [self.first_col] + self.cols
        c = 1
        while c <= len(allcols):
            g = allcols[c - 1].group
            span = 1
            while g and c + span <= len(allcols) and allcols[c + span - 1].group == g:
                span += 1
            cell = ws.cell(row=g_row, column=c, value=g or None)
            cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=BLUE if g else NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            white = Side(style="thin", color="FFFFFF")
            cell.border = Border(left=white, right=white)
            for k in range(1, span):
                ws.cell(row=g_row, column=c + k).fill = PatternFill(
                    "solid", fgColor=BLUE
                )
            if span > 1:
                ws.merge_cells(
                    start_row=g_row,
                    start_column=c,
                    end_row=g_row,
                    end_column=c + span - 1,
                )
            c += span
        ws.row_dimensions[g_row].height = 16
        for i, col in enumerate(allcols, 1):
            cell = ws.cell(row=h_row, column=i, value=col.label)
            cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(
                horizontal="center" if i > 1 else "left",
                vertical="center",
                wrap_text=True,
            )
            ws.column_dimensions[L(i)].width = col.width
        ws.row_dimensions[h_row].height = 44
        ws.freeze_panes = ws.cell(row=h_row + 1, column=2)

    def next_row(self) -> int:
        return self.header_row + 1 + self.rows_written

    def row(self, values: list, *, style: str = "data") -> int:
        r = self.next_row()
        self.rows_written += 1
        allcols = [self.first_col] + self.cols
        band = self.rows_written % 2 == 0
        for i, v in enumerate(values, 1):
            cell = self.ws.cell(row=r, column=i, value=v)
            col = allcols[i - 1]
            bold = style in ("total", "subtotal")
            cell.font = Font(
                name=FONT,
                size=9 if style == "note" else 10,
                bold=bold,
                italic=style == "note",
                color=GREY if style == "note" else "000000",
            )
            cell.number_format = col.fmt if i > 1 else "@"
            cell.alignment = Alignment(
                horizontal="right" if i > 1 else "left", vertical="center"
            )
            cell.border = Border(bottom=THIN)
            if style == "total":
                cell.fill = PatternFill("solid", fgColor=SUBTOTAL)
            elif style == "subtotal":
                cell.fill = PatternFill("solid", fgColor="EEF2F8")
            elif band and style == "data":
                cell.fill = PatternFill("solid", fgColor=BAND)
        self.ws.row_dimensions[r].height = 15
        return r

    def finish(self, data_last: int | None = None) -> None:
        ws = self.ws
        first = self.header_row + 1
        last = self.header_row + self.rows_written
        data_last = data_last or last
        ws.auto_filter.ref = f"A{self.header_row}:{L(len(self.cols) + 1)}{data_last}"
        for i, col in enumerate(self.cols, 2):
            if col.scale and data_last >= first:
                ref = f"{L(i)}{first}:{L(i)}{data_last}"
                ws.conditional_formatting.add(
                    ref,
                    ColorScaleRule(
                        start_type="min",
                        start_color="F8CBAD",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFF2CC",
                        end_type="max",
                        end_color="C6E0B4",
                    ),
                )
        ws.print_title_rows = f"{self.header_row - 1}:{self.header_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True


def pct(num: str, den: str) -> str:
    return f'=IF(AND(ISNUMBER({num}),ISNUMBER({den}),{den}>0),{num}/{den},"")'


def growth(new: str, old: str) -> str:
    return f'=IF(AND(ISNUMBER({new}),ISNUMBER({old}),{old}>0),{new}/{old}-1,"")'


def cagr(new: str, old: str, years: int) -> str:
    return f'=IF(AND(ISNUMBER({new}),ISNUMBER({old}),{old}>0),({new}/{old})^(1/{years})-1,"")'


# ---------------------------------------------------------------- build


def build(raw: dict, areas: dict, series: dict, out: Path) -> None:
    county = raw["county"]["Demography"]
    subs = raw["county"]["SubRegionals"]
    quarters = raw["quarters"]
    el_q = raw["election"]["quarters"]
    fel_q = raw["fellows"]["quarters"]
    order = sorted(subs, key=lambda s: (s["DistrictId"] >= 100000, s["RegionName"]))
    rows = []
    for s in order:
        did = str(s["DistrictId"])
        q = quarters.get(did)
        if not q:
            continue
        area = areas.get(norm_name(s["RegionName"]))
        rows.append(
            {
                "did": did,
                "name": s["RegionName"],
                "q": q,
                "area_id": area["area_id"] if area else "",
                "kind": "centre" if s["DistrictId"] < 100000 else "rural",
                "ph": q["HouseholdCount"] == 0,
                "tuik": series.get(area["area_id"], {}) if area else {},
            }
        )
    n = len(rows)
    place = f"{county['CityName'].title()} {county['CountyName'].title()}"
    fetched = raw["county"].get("_meta", {}).get("fetched", "")
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True

    R0 = 5  # header_row 4 → first data row 5; same row index on every per-neighbourhood sheet

    def rr(i: int) -> int:
        return R0 + i

    R_CENTRE, R_RURAL, R_COUNTY = R0 + n, R0 + n + 1, R0 + n + 2
    centre_rows = [rr(i) for i, r in enumerate(rows) if r["kind"] == "centre"]
    rural_rows = [rr(i) for i, r in enumerate(rows) if r["kind"] == "rural"]
    yrs = sorted({y for s in series.values() for y in s})
    Y = len(yrs)

    def sum_rows(col: str, idx: list[int]) -> str | None:
        if not idx:
            return None
        if idx == list(range(idx[0], idx[-1] + 1)):
            return f"=SUM({col}{idx[0]}:{col}{idx[-1]})"
        return "=" + "+".join(f"{col}{r}" for r in idx)

    def add_subtotals(
        t: Table,
        numeric_cols: list[int],
        county_vals: dict[int, object] | None = None,
        formula_cols: dict[int, str] | None = None,
        control: bool = True,
    ) -> None:
        """Append MERKEZ / KIR / İLÇE rows (+ control sum). formula_cols: col -> template with {r}."""
        ncol = len(t.cols) + 1
        for label, idx, style in (
            ("MERKEZ toplamı", centre_rows, "subtotal"),
            ("KIR toplamı", rural_rows, "subtotal"),
        ):
            vals: list = [label] + [None] * (ncol - 1)
            r = t.next_row()
            for c in numeric_cols:
                vals[c - 1] = sum_rows(L(c), idx)
            for c, f in (formula_cols or {}).items():
                vals[c - 1] = f.format(r=r)
            t.row(vals, style=style)
        vals = ["İLÇE (Endeksa)"] + [None] * (ncol - 1)
        r = t.next_row()
        for c, v in (county_vals or {}).items():
            vals[c - 1] = v
        for c, f in (formula_cols or {}).items():
            vals[c - 1] = f.format(r=r)
        t.row(vals, style="total")
        if control:
            vals = ["Mahalleler toplamı (kontrol)"] + [None] * (ncol - 1)
            for c in numeric_cols:
                vals[c - 1] = f"=SUM({L(c)}{R0}:{L(c)}{R0 + n - 1})"
            t.row(vals, style="note")

    YT, YE, YK = "'Yaş Toplam'", "'Yaş Erkek'", "'Yaş Kadın'"
    mids = "{" + ",".join(str(m) for m in AGE_MID) + "}"

    # ================================================================ Kapak
    ws = wb.active
    ws.title = "Kapak"
    ws.sheet_view.showGridLines = False
    ws["B2"] = f"{place} — mahalle veri kitabı"
    ws["B2"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
    ws["B3"] = (
        f"Endeksa 2024 kesiti (döküm {fetched}) · TÜİK ADNKS {yrs[0]}-{yrs[-1]} · VeriAtlas"
    )
    ws["B3"].font = Font(name=FONT, italic=True, size=10, color=GREY)
    kpis = [
        ("Nüfus 2024", county["PopulationTotal"], FMT_INT),
        ("Yüzölçümü km²", county["Area"], FMT_DEC1),
        ("Mahalle", n, FMT_INT),
        ("Veri olan mahalle", sum(1 for r in rows if not r["ph"]), FMT_INT),
        ("65+ payı", county["Age_65_Total"] / county["PopulationTotal"], FMT_PCT),
        (
            "Ortalama yaş",
            sum(county[f"Age_{b}_Total"] * m for b, m in zip(AGE_BANDS, AGE_MID))
            / county["PopulationTotal"],
            FMT_DEC1,
        ),
        (
            "Hane büyüklüğü",
            county["PopulationTotal"] / county["HouseholdCount"],
            FMT_DEC2,
        ),
    ]
    for i, (k, v, f) in enumerate(kpis):
        c = 2 + i
        ws.cell(row=5, column=c, value=k).font = Font(name=FONT, size=9, color=GREY)
        cell = ws.cell(row=6, column=c, value=v)
        cell.font = Font(name=FONT, bold=True, size=16, color=NAVY)
        cell.number_format = f
        ws.column_dimensions[L(c)].width = 17
    ws["B9"] = "Sayfalar"
    ws["B9"].font = Font(name=FONT, bold=True, size=11, color=NAVY)
    index = [
        ("Özet", "Mahalle başına temel göstergeler"),
        (
            "Analiz",
            "Ortalama/medyan yaş, bağımlılık, yaşlanma endeksi; merkez-kır-ilçe",
        ),
        ("Kimlik", "MEDAS ve Endeksa kimlikleri, tür, veri ve sınır durumu"),
        ("Yaş Toplam / Erkek / Kadın", "5'lik yaş grupları, kişi"),
        ("Yaş Payları", "Yaş grubu payları ve cinsiyet oranı"),
        ("Eğitim", "Eğitim düzeyi kişi ve pay"),
        ("Medeni Hal", "15+ medeni hal kişi ve pay"),
        ("SES", "Sosyo-ekonomik statü grupları (Endeksa modeli)"),
        ("Gelir-Harcama", "Hane geliri, tasarruf, harcama kalemleri, mülkiyet"),
        ("Konut", "Konut stoğu, ticari birim, m² fiyat"),
        ("Emlak Satış", "Tapu satış ve ilan serileri 2012-2024"),
        ("Seçim 2024", "2024 ilçe belediye başkanlığı, mahalle × parti"),
        ("Katılım", "17 seçimde katılım, mahalle × seçim"),
        ("Seçimler", "Tüm seçimler, uzun tablo (pivot için)"),
        ("Hemşehri", "Nüfusa kayıtlı il, ilk 10"),
        ("TÜİK Nüfus", f"Yıllık toplam nüfus {yrs[0]}-{yrs[-1]}"),
        ("TÜİK Yaş-Cinsiyet", "Yıllık 18+ / 0-17 × erkek / kadın"),
        ("Ham", "Endeksa demografi yanıtının tüm alanları"),
        ("Notlar", "Kaynak, yıl, sınırlar, yöntem"),
    ]
    for i, (k, v) in enumerate(index, 10):
        ws.cell(row=i, column=2, value=k).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=i, column=3, value=v).font = Font(name=FONT, size=10)
    r0 = 10 + len(index) + 1
    ws.cell(row=r0, column=2, value="Renkler").font = Font(
        name=FONT, bold=True, size=11, color=NAVY
    )
    legend = [
        ("Açık mavi satır", "Merkez / Kır ara toplamı", "EEF2F8"),
        ("Koyu mavi satır", "İlçe (Endeksa)", SUBTOTAL),
        (
            "Turuncu → yeşil",
            "Sütun içinde düşükten yükseğe (yalnız oran sütunları)",
            "FFF2CC",
        ),
        (
            "(veri yok)",
            "Endeksa mahalle demografisi boş; TÜİK ve seçim geçerli",
            "FFFFFF",
        ),
    ]
    for i, (k, v, col) in enumerate(legend, r0 + 1):
        c = ws.cell(row=i, column=2, value=k)
        c.fill = PatternFill("solid", fgColor=col)
        c.font = Font(name=FONT, size=10)
        ws.cell(row=i, column=3, value=v).font = Font(name=FONT, size=10)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 17

    # ================================================================ Yaş × cinsiyet
    for sfx, lbl in SEX:
        t = Table(
            wb.create_sheet(f"Yaş {lbl}"),
            f"{place} — yaş grupları, {lbl.lower()} (kişi)",
            "Endeksa 2024 (TÜİK ADNKS kökenli). Toplam sütunu formül.",
            [Col("Toplam", FMT_INT, 9)]
            + [Col(AGE_LABEL[b], FMT_INT, 7, group="Yaş grubu") for b in AGE_BANDS],
        )
        for i, r in enumerate(rows):
            R = rr(i)
            if r["ph"]:
                t.row([r["name"] + " (veri yok)"] + [None] * 15)
                continue
            t.row(
                [r["name"], f"=SUM(C{R}:P{R})"]
                + [r["q"][f"Age_{b}_{sfx}"] for b in AGE_BANDS]
            )
        add_subtotals(
            t,
            list(range(3, 17)),
            {c: county[f"Age_{b}_{sfx}"] for c, b in zip(range(3, 17), AGE_BANDS)},
            {2: "=SUM(C{r}:P{r})"},
        )
        t.finish(R0 + n - 1)

    # ================================================================ Yaş payları
    t = Table(
        wb.create_sheet("Yaş Payları"),
        f"{place} — yaş grubu payları ve cinsiyet oranı",
        "Formül: 'Yaş Toplam' sayfasına bölünerek. Cinsiyet oranı = erkek / kadın × 100.",
        [
            Col(g, FMT_PCT, 8, True, "Geniş yaş grubu payı")
            for g in ("0-14", "15-29", "30-44", "45-64", "65+")
        ]
        + [Col(AGE_LABEL[b], FMT_PCT, 7, group="5'lik grup payı") for b in AGE_BANDS]
        + [
            Col("Kadın payı", FMT_PCT, 8, True, "Cinsiyet"),
            Col("Erkek / 100 kadın", "0", 9, True, "Cinsiyet"),
            Col("65+ erkek / 100 kadın", "0", 10, True, "Cinsiyet"),
        ],
    )

    def share_vals(R: int) -> list:
        tot = f"{YT}!B{R}"
        v = [
            f'=IF({tot}>0,SUM({YT}!C{R}:E{R})/{tot},"")',
            f'=IF({tot}>0,SUM({YT}!F{R}:H{R})/{tot},"")',
            f'=IF({tot}>0,SUM({YT}!I{R}:K{R})/{tot},"")',
            f'=IF({tot}>0,SUM({YT}!L{R}:O{R})/{tot},"")',
            f'=IF({tot}>0,{YT}!P{R}/{tot},"")',
        ]
        v += [f'=IF({tot}>0,{YT}!{L(c)}{R}/{tot},"")' for c in range(3, 17)]
        v += [
            f'=IF({tot}>0,{YK}!B{R}/{tot},"")',
            f'=IF({YK}!B{R}>0,{YE}!B{R}/{YK}!B{R}*100,"")',
            f'=IF({YK}!P{R}>0,{YE}!P{R}/{YK}!P{R}*100,"")',
        ]
        return v

    for i, r in enumerate(rows):
        t.row(
            [r["name"] + " (veri yok)"] + [None] * 22
            if r["ph"]
            else [r["name"]] + share_vals(rr(i))
        )
    for R, label, style in (
        (R_CENTRE, "MERKEZ", "subtotal"),
        (R_RURAL, "KIR", "subtotal"),
        (R_COUNTY, "İLÇE (Endeksa)", "total"),
    ):
        t.row([label] + share_vals(R), style=style)
    t.finish(R0 + n - 1)

    # ================================================================ Analiz
    t = Table(
        wb.create_sheet("Analiz"),
        f"{place} — yaş yapısı ve nüfus dinamiği",
        "Ortalama yaş: 5'lik grup orta noktaları (65+ için 74) ile ağırlıklı ortalama, formül. Medyan: grup içi doğrusal ara değer (Python). Bağımlılık oranları 15-64'e göre. Değişim TÜİK.",
        [
            Col("Ortalama yaş", FMT_DEC1, 9, True, "Yaş"),
            Col("Medyan yaş", FMT_DEC1, 9, True, "Yaş"),
            Col("Ort. yaş erkek", FMT_DEC1, 9, group="Yaş"),
            Col("Ort. yaş kadın", FMT_DEC1, 9, group="Yaş"),
            Col("Genç (0-14 / 15-64)", FMT_PCT, 10, True, "Bağımlılık oranı"),
            Col("Yaşlı (65+ / 15-64)", FMT_PCT, 10, True, "Bağımlılık oranı"),
            Col("Toplam", FMT_PCT, 9, True, "Bağımlılık oranı"),
            Col("Yaşlanma endeksi (65+ / 0-14 × 100)", "0", 12, True, "Yapı"),
            Col("Hane büyüklüğü", FMT_DEC2, 9, True, "Yapı"),
            Col(f"Nüfus {yrs[0]}", FMT_INT, 9, group="TÜİK"),
            Col(f"Nüfus {yrs[-1]}", FMT_INT, 9, group="TÜİK"),
            Col(f"Değişim {yrs[0]}→{yrs[-1]}", FMT_PCT, 10, True, "TÜİK"),
            Col("Yıllık ort. (CAGR)", "0.00%", 9, True, "TÜİK"),
        ],
    )

    def analysis_vals(
        R: int, counts: list[int] | None, hh: int | None, tuik_first, tuik_last
    ) -> list:
        tot = f"{YT}!B{R}"
        c15_64 = f"SUM({YT}!F{R}:O{R})"
        c0_14 = f"SUM({YT}!C{R}:E{R})"
        return [
            f'=IF({tot}>0,SUMPRODUCT({YT}!C{R}:P{R},{mids})/{tot},"")',
            median_age(counts) if counts else None,
            f'=IF({YE}!B{R}>0,SUMPRODUCT({YE}!C{R}:P{R},{mids})/{YE}!B{R},"")',
            f'=IF({YK}!B{R}>0,SUMPRODUCT({YK}!C{R}:P{R},{mids})/{YK}!B{R},"")',
            f'=IF({c15_64}>0,{c0_14}/{c15_64},"")',
            f'=IF({c15_64}>0,{YT}!P{R}/{c15_64},"")',
            f'=IF({c15_64}>0,({c0_14}+{YT}!P{R})/{c15_64},"")',
            f'=IF({c0_14}>0,{YT}!P{R}/{c0_14}*100,"")',
            f'=IF({hh or 0}>0,{tot}/{hh},"")' if hh else None,
            tuik_first,
            tuik_last,
            growth(f"L{R}", f"K{R}"),
            cagr(f"L{R}", f"K{R}", Y - 1),
        ]

    for i, r in enumerate(rows):
        R = rr(i)
        tf, tl = tuik_get(r["tuik"], yrs[0]), tuik_get(r["tuik"], yrs[-1])
        if r["ph"]:
            t.row(
                [r["name"] + " (veri yok)"]
                + [None] * 9
                + [tf, tl, growth(f"L{R}", f"K{R}"), cagr(f"L{R}", f"K{R}", Y - 1)]
            )
            continue
        q = r["q"]
        t.row(
            [r["name"]]
            + analysis_vals(
                R, [q[f"Age_{b}_Total"] for b in AGE_BANDS], q["HouseholdCount"], tf, tl
            )
        )
    for R, label, idx in (
        (R_CENTRE, "MERKEZ", centre_rows),
        (R_RURAL, "KIR", rural_rows),
    ):
        sel = [rows[x - R0] for x in idx]
        counts = [
            sum(s["q"][f"Age_{b}_Total"] for s in sel if not s["ph"]) for b in AGE_BANDS
        ]
        hh = sum(s["q"]["HouseholdCount"] for s in sel)
        t.row(
            [label]
            + analysis_vals(R, counts, hh, sum_rows("K", idx), sum_rows("L", idx)),
            style="subtotal",
        )
    t.row(
        ["İLÇE (Endeksa)"]
        + analysis_vals(
            R_COUNTY,
            [county[f"Age_{b}_Total"] for b in AGE_BANDS],
            county["HouseholdCount"],
            None,
            None,
        ),
        style="total",
    )
    t.finish(R0 + n - 1)

    # ================================================================ Eğitim
    t = Table(
        wb.create_sheet("Eğitim"),
        f"{place} — eğitim düzeyi (6+ yaş)",
        "Kişi sayıları Endeksa 2024 (TÜİK kökenli); toplam ve paylar formül. Alt = bilmeyen + bitirmemiş + ilkokul; Orta = ilköğretim + ortaokul + lise; Yüksek = lisans + YL + doktora.",
        [Col(lbl, FMT_INT, 9, group="Kişi") for _, lbl in EDU]
        + [
            Col("Toplam", FMT_INT, 9, group="Kişi"),
            Col("Alt", FMT_PCT, 8, True, "Pay"),
            Col("Orta", FMT_PCT, 8, True, "Pay"),
            Col("Yüksek", FMT_PCT, 8, True, "Pay"),
            Col("Okuma yazma bilmeyen", FMT_PCT, 9, True, "Pay"),
        ],
    )

    def edu_vals(q, R):
        return [q[k] for k, _ in EDU] + [
            f"=SUM(B{R}:K{R})",
            pct(f"SUM(B{R}:D{R})", f"L{R}"),
            pct(f"SUM(E{R}:G{R})", f"L{R}"),
            pct(f"SUM(H{R}:J{R})", f"L{R}"),
            pct(f"B{R}", f"L{R}"),
        ]

    for i, r in enumerate(rows):
        t.row(
            [r["name"] + " (veri yok)"] + [None] * 15
            if r["ph"]
            else [r["name"]] + edu_vals(r["q"], rr(i))
        )
    add_subtotals(
        t,
        list(range(2, 12)),
        {c + 2: county[k] for c, (k, _) in enumerate(EDU)},
        {
            12: "=SUM(B{r}:K{r})",
            13: '=IF(L{r}>0,SUM(B{r}:D{r})/L{r},"")',
            14: '=IF(L{r}>0,SUM(E{r}:G{r})/L{r},"")',
            15: '=IF(L{r}>0,SUM(H{r}:J{r})/L{r},"")',
            16: '=IF(L{r}>0,B{r}/L{r},"")',
        },
    )
    t.finish(R0 + n - 1)

    # ================================================================ Medeni
    t = Table(
        wb.create_sheet("Medeni Hal"),
        f"{place} — medeni hal (15+ yaş)",
        "Kişi sayıları Endeksa 2024 (TÜİK kökenli); toplam ve paylar formül.",
        [Col(lbl, FMT_INT, 10, group="Kişi") for _, lbl in MARITAL]
        + [Col("Toplam 15+", FMT_INT, 10, group="Kişi")]
        + [Col(lbl, FMT_PCT, 10, True, "Pay") for _, lbl in MARITAL],
    )
    for i, r in enumerate(rows):
        R = rr(i)
        t.row(
            [r["name"] + " (veri yok)"] + [None] * 9
            if r["ph"]
            else [r["name"]]
            + [r["q"][k] for k, _ in MARITAL]
            + [f"=SUM(B{R}:E{R})"]
            + [pct(f"{L(c)}{R}", f"F{R}") for c in range(2, 6)]
        )
    add_subtotals(
        t,
        [2, 3, 4, 5],
        {c + 2: county[k] for c, (k, _) in enumerate(MARITAL)},
        {
            6: "=SUM(B{r}:E{r})",
            **{6 + j: f'=IF(F{{r}}>0,{L(1 + j)}{{r}}/F{{r}},"")' for j in range(1, 5)},
        },
    )
    t.finish(R0 + n - 1)

    # ================================================================ SES
    t = Table(
        wb.create_sheet("SES"),
        f"{place} — sosyo-ekonomik statü",
        "Endeksa modeli (tahmin; TÜİK sayımı değil). Kişi sayıları kaynaktan, paylar formül. Endeks etiketi: Endeksa'nın mahalleyi Türkiye / il içinde konumlandırması.",
        [Col(lbl, FMT_INT, 8, group="Kişi") for _, lbl in SES]
        + [
            Col("Toplam", FMT_INT, 9, group="Kişi"),
            Col("A+ A B", FMT_PCT, 8, True, "Pay"),
            Col("C", FMT_PCT, 8, True, "Pay"),
            Col("D", FMT_PCT, 8, True, "Pay"),
            Col("Türkiye", "@", 11, group="Endeks"),
            Col("İl", "@", 11, group="Endeks"),
        ],
    )
    for i, r in enumerate(rows):
        R = rr(i)
        q = r["q"]
        t.row(
            [r["name"] + " (veri yok)"] + [None] * 11
            if r["ph"]
            else [r["name"]]
            + [q[k] for k, _ in SES]
            + [
                f"=SUM(B{R}:F{R})",
                pct(f"SUM(B{R}:D{R})", f"G{R}"),
                pct(f"E{R}", f"G{R}"),
                pct(f"F{R}", f"G{R}"),
                q.get("TurkeyIndex"),
                q.get("CityIndex"),
            ]
        )
    add_subtotals(
        t,
        list(range(2, 7)),
        {c + 2: county[k] for c, (k, _) in enumerate(SES)},
        {
            7: "=SUM(B{r}:F{r})",
            8: '=IF(G{r}>0,SUM(B{r}:D{r})/G{r},"")',
            9: '=IF(G{r}>0,E{r}/G{r},"")',
            10: '=IF(G{r}>0,F{r}/G{r},"")',
        },
    )
    t.finish(R0 + n - 1)

    # ================================================================ Gelir-Harcama
    t = Table(
        wb.create_sheet("Gelir-Harcama"),
        f"{place} — gelir, tasarruf, harcama (₺/ay) ve konut mülkiyeti",
        "Endeksa modeli (tahmin). Harcama payları formül (kalem / harcama toplamı).",
        [
            Col("Hane geliri", FMT_INT, 10, True, "Gelir"),
            Col("Kişi başı", FMT_INT, 9, True, "Gelir"),
            Col("Tasarruf", FMT_INT, 9, group="Gelir"),
            Col("Harcama", FMT_INT, 9, group="Gelir"),
        ]
        + [Col(lbl, FMT_INT, 8, group="Harcama kalemi ₺") for _, lbl in EXPENSE]
        + [
            Col(g, FMT_PCT, 8, True, "Harcama payı")
            for g in ("Gıda", "Barınma", "Ulaşım", "Eğitim")
        ]
        + [
            Col("Mülk sahibi", FMT_PCT, 9, True, "Mülkiyet"),
            Col("Kiracı", FMT_PCT, 9, True, "Mülkiyet"),
            Col("GSYH ₺", FMT_INT, 13, group="Diğer"),
            Col("Araç", FMT_INT, 8, group="Diğer"),
            Col("Mobil kullanıcı", FMT_INT, 9, group="Diğer"),
        ],
    )

    def inc_vals(
        q, R
    ):  # B gelir, C kişi başı, D tasarruf, E harcama, F..Q kalemler (gıda F, barınma I, ulaşım L, eğitim O)
        return (
            [
                q["HouseIncomeTotal"],
                q["HouseIncome"],
                q["SavingTotal"],
                q["ExpenseTotal"],
            ]
            + [q[k] for k, _ in EXPENSE]
            + [
                pct(f"F{R}", f"E{R}"),
                pct(f"I{R}", f"E{R}"),
                pct(f"L{R}", f"E{R}"),
                pct(f"O{R}", f"E{R}"),
                q["OwnerShare"] / 100,
                q["RentedShare"] / 100,
                q["GSYH"],
                q["CarCount"],
                q["MobileUser"],
            ]
        )

    for i, r in enumerate(rows):
        t.row(
            [r["name"] + " (veri yok)"] + [None] * 25
            if r["ph"]
            else [r["name"]] + inc_vals(r["q"], rr(i))
        )
    t.row(["İLÇE (Endeksa)"] + inc_vals(county, R0 + n), style="total")
    t.finish(R0 + n - 1)

    # ================================================================ Konut
    t = Table(
        wb.create_sheet("Konut"),
        f"{place} — konut stoğu ve fiyatlar",
        "Endeksa. Konut / hane formül. Fiyatlar ilan bazlı m² (₺); boş = ilan yok.",
        [
            Col("Konut", FMT_INT, 9, group="Stok (adet)"),
            Col("Yazlık", FMT_INT, 8, group="Stok (adet)"),
            Col("Ticari birim", FMT_INT, 9, group="Stok (adet)"),
            Col("Hane", FMT_INT, 9, group="Stok (adet)"),
            Col("Konut / hane", FMT_DEC2, 9, True, "Stok (adet)"),
            Col("Konut satış", FMT_INT, 9, group="m² fiyat ₺"),
            Col("Konut kira", FMT_INT, 9, group="m² fiyat ₺"),
            Col("Ticari satış", FMT_INT, 9, group="m² fiyat ₺"),
            Col("Ticari kira", FMT_INT, 9, group="m² fiyat ₺"),
            Col("Arsa", FMT_INT, 8, group="m² fiyat ₺"),
            Col("Tarla", FMT_INT, 8, group="m² fiyat ₺"),
            Col("Konut satış (gün)", FMT_INT, 9, group="İlan süresi"),
            Col("Konut kira (gün)", FMT_INT, 9, group="İlan süresi"),
        ],
    )

    def house_vals(q, R):
        return [
            q["HousingCount"],
            q["SummerResortCount"],
            q["CommercialCount"],
            q["HouseholdCount"] or None,
            f'=IF(AND(ISNUMBER(E{R}),E{R}>0),B{R}/E{R},"")',
        ] + [
            q[k] or None
            for k in (
                "HouseUnitPriceForSale",
                "HouseUnitPriceForRent",
                "CommercialUnitPriceForSale",
                "CommercialUnitPriceForRent",
                "PlotUnitPriceForSale",
                "LandUnitPriceForSale",
                "HouseListingPeriodForSale",
                "HouseListingPeriodForRent",
            )
        ]

    for i, r in enumerate(rows):
        t.row([r["name"]] + house_vals(r["q"], rr(i)))
    add_subtotals(
        t,
        [2, 3, 4, 5],
        dict(zip(range(2, 15), house_vals(county, R_COUNTY))),
        {6: '=IF(AND(ISNUMBER(E{r}),E{r}>0),B{r}/E{r},"")'},
    )
    t.finish(R0 + n - 1)

    # ================================================================ Emlak satış
    years = list(range(2012, 2025))
    lyears = list(range(2014, 2025))
    t = Table(
        wb.create_sheet("Emlak Satış"),
        f"{place} — tapu satışları ve ilan sayıları (adet / yıl)",
        "Endeksa (tapu + ilan). 2024 kısmi yıl. Toplam sütunları formül.",
        [Col(str(y), FMT_INT, 6.5, group="Konut satışı") for y in years]
        + [Col("Toplam", FMT_INT, 8, group="Konut satışı")]
        + [Col(str(y), FMT_INT, 6.5, group="Arsa-tarla satışı") for y in years]
        + [Col("Toplam", FMT_INT, 8, group="Arsa-tarla satışı")]
        + [Col(str(y), FMT_INT, 6.5, group="İlan") for y in lyears],
    )

    def sale_vals(q, R):
        return (
            [q.get(f"Total_BB_Sale_{y}", 0) for y in years]
            + [f"=SUM(B{R}:N{R})"]
            + [q.get(f"Total_AT_Sale_{y}", 0) for y in years]
            + [f"=SUM(P{R}:AB{R})"]
            + [q.get(f"Total_Listing_{y}", 0) for y in lyears]
        )

    for i, r in enumerate(rows):
        t.row([r["name"]] + sale_vals(r["q"], rr(i)))
    ncols = 2 * 14 + len(lyears)
    add_subtotals(
        t,
        [c for c in range(2, ncols + 2) if c not in (15, 29)],
        dict(zip(range(2, ncols + 2), sale_vals(county, R_COUNTY))),
        {15: "=SUM(B{r}:N{r})", 29: "=SUM(P{r}:AB{r})"},
    )
    t.finish(R0 + n - 1)

    # ================================================================ Seçim 2024
    parties: list[str] = []
    for lst in el_q.values():
        e = next((x for x in lst if x["Code"] == FOCUS_ELECTION), None)
        for s in e["Secenekler"] if e else []:
            if s["Secenek"] not in parties:
                parties.append(s["Secenek"])
    ec = next(
        (x for x in raw["election"]["county"] if x["Code"] == FOCUS_ELECTION), None
    )
    cv = {s["Secenek"]: s["OySayisi"] for s in ec["Secenekler"]} if ec else {}
    parties.sort(key=lambda p: -cv.get(p, 0))
    # keep parties above 1% of county valid votes; fold the rest into "Diğer"
    threshold = 0.01 * (ec["GecerliOy"] if ec else 0)
    minor = {p for p in parties if cv.get(p, 0) < threshold}
    parties = [p for p in parties if p not in minor and p != "Diğer"] + ["Diğer"]
    P = len(parties)
    VOTE0, SHARE0 = 8, 8 + P
    TOPC, TOPS, MARGIN = SHARE0 + P, SHARE0 + P + 1, SHARE0 + P + 2
    t = Table(
        wb.create_sheet("Seçim 2024"),
        f"{place} — {ELECTION_LABEL[FOCUS_ELECTION]} başkanlığı, mahalle × parti",
        "Oy sayıları Endeksa (YSK). Katılım = kullanılan / kayıtlı; pay = oy / geçerli; formül. İlçe geçerli oyunun %1'i altındaki partiler ve kaynaktaki 'Diğer' tek sütunda toplandı (tam döküm 'Seçimler' sayfasında).",
        [
            Col("Sandık", "0", 7, group="Seçmen"),
            Col("Kayıtlı", FMT_INT, 9, group="Seçmen"),
            Col("Kullanılan", FMT_INT, 9, group="Seçmen"),
            Col("Geçerli", FMT_INT, 9, group="Seçmen"),
            Col("Geçersiz", FMT_INT, 8, group="Seçmen"),
            Col("Katılım", FMT_PCT, 8, True, "Seçmen"),
        ]
        + [Col(p, FMT_INT, 8, group="Oy") for p in parties]
        + [Col(p, FMT_PCT, 8, True, "Pay") for p in parties]
        + [
            Col("1. parti", "@", 10, group="Sonuç"),
            Col("1. parti payı", FMT_PCT, 8, True, "Sonuç"),
            Col("Fark 1.-2. (puan)", "0.0", 9, True, "Sonuç"),
        ],
    )
    hdr = f"${L(VOTE0)}$4:${L(VOTE0 + P - 1)}$4"

    def el_vals(e, R):
        if not e:
            return [None] * (6 + 2 * P + 3)
        votes: dict[str, float] = {}
        for s in e["Secenekler"]:
            key = (
                "Diğer"
                if s["Secenek"] in minor or s["Secenek"] == "Diğer"
                else s["Secenek"]
            )
            votes[key] = votes.get(key, 0) + s["OySayisi"]
        vr = f"{L(VOTE0)}{R}:{L(VOTE0 + P - 1)}{R}"
        sr = f"{L(SHARE0)}{R}:{L(SHARE0 + P - 1)}{R}"
        return (
            [
                e["SandikSayisi"],
                e["KayitliSecmen"],
                e["KullanilanOy"],
                e["GecerliOy"],
                e["GecersizOy"],
                pct(f"D{R}", f"C{R}"),
            ]
            + [votes.get(p, 0) for p in parties]
            + [pct(f"{L(VOTE0 + j)}{R}", f"E{R}") for j in range(P)]
            + [
                f'=IF(E{R}>0,INDEX({hdr},MATCH(MAX({vr}),{vr},0)),"")',
                f'=IF(E{R}>0,MAX({sr}),"")',
                f'=IF(E{R}>0,(LARGE({sr},1)-LARGE({sr},2))*100,"")',
            ]
        )

    for i, r in enumerate(rows):
        e = next(
            (x for x in el_q.get(r["did"], []) if x["Code"] == FOCUS_ELECTION), None
        )
        t.row([r["name"]] + el_vals(e, rr(i)))
    sub_formulas = {
        7: '=IF(C{r}>0,D{r}/C{r},"")',
        **{
            SHARE0 + j: f'=IF(E{{r}}>0,{L(VOTE0 + j)}{{r}}/E{{r}},"")' for j in range(P)
        },
    }
    sub_formulas[TOPC] = (
        f'=IF(E{{r}}>0,INDEX({hdr},MATCH(MAX({L(VOTE0)}{{r}}:{L(VOTE0 + P - 1)}{{r}}),{L(VOTE0)}{{r}}:{L(VOTE0 + P - 1)}{{r}},0)),"")'
    )
    sub_formulas[TOPS] = (
        f'=IF(E{{r}}>0,MAX({L(SHARE0)}{{r}}:{L(SHARE0 + P - 1)}{{r}}),"")'
    )
    sub_formulas[MARGIN] = (
        f'=IF(E{{r}}>0,(LARGE({L(SHARE0)}{{r}}:{L(SHARE0 + P - 1)}{{r}},1)-LARGE({L(SHARE0)}{{r}}:{L(SHARE0 + P - 1)}{{r}},2))*100,"")'
    )
    add_subtotals(
        t,
        [2, 3, 4, 5, 6] + list(range(VOTE0, VOTE0 + P)),
        dict(zip(range(2, 2 + 6 + 2 * P + 3), el_vals(ec, R_COUNTY))),
        sub_formulas,
    )
    t.finish(R0 + n - 1)

    # ================================================================ Katılım
    codes = [e["Code"] for e in raw["election"]["county"]]
    C = len(codes)
    t = Table(
        wb.create_sheet("Katılım"),
        f"{place} — seçim katılımı, mahalle × seçim",
        "Katılım = kullanılan / kayıtlı (formül; sayılar 'Seçimler' sayfasında). Kaynak: Endeksa (YSK).",
        [
            Col(ELECTION_LABEL.get(c, c), FMT_PCT, 8.5, True, "Katılım oranı")
            for c in codes
        ]
        + [
            Col("2024", FMT_INT, 9, group="Kayıtlı seçmen"),
            Col("2011", FMT_INT, 9, group="Kayıtlı seçmen"),
            Col("Değişim", FMT_PCT, 9, True, "Kayıtlı seçmen"),
        ],
    )

    def turnout_vals(by: dict, R: int) -> list:
        v = [
            f"={by[c]['KullanilanOy']}/{by[c]['KayitliSecmen']}"
            if c in by and by[c]["KayitliSecmen"]
            else None
            for c in codes
        ]
        k24 = by.get(FOCUS_ELECTION, {}).get("KayitliSecmen")
        k11 = by.get("2011genelsecim", {}).get("KayitliSecmen")
        return v + [k24, k11, growth(f"{L(C + 2)}{R}", f"{L(C + 3)}{R}")]

    for i, r in enumerate(rows):
        t.row(
            [r["name"]]
            + turnout_vals({e["Code"]: e for e in el_q.get(r["did"], [])}, rr(i))
        )
    t.row(
        ["İLÇE"]
        + turnout_vals({e["Code"]: e for e in raw["election"]["county"]}, R0 + n),
        style="total",
    )
    t.finish(R0 + n - 1)

    # ================================================================ Seçimler (long)
    t = Table(
        wb.create_sheet("Seçimler"),
        f"{place} — tüm seçimler, uzun tablo",
        "Her satır bir mahalle × seçim × seçenek; özet tablo (pivot) için. Pay = oy / geçerli (formül).",
        [
            Col("Tür", "@", 8),
            Col("Seçim", "@", 18),
            Col("Seçim kodu", "@", 30),
            Col("Sandık", "0", 7),
            Col("Kayıtlı", FMT_INT, 9),
            Col("Kullanılan", FMT_INT, 9),
            Col("Geçerli", FMT_INT, 9),
            Col("Geçersiz", FMT_INT, 8),
            Col("Seçenek", "@", 22),
            Col("Oy", FMT_INT, 9),
            Col("Pay", FMT_PCT, 8),
        ],
    )
    for r in rows + [{"name": "İLÇE", "did": "__county__", "kind": None}]:
        lst = (
            raw["election"]["county"]
            if r["did"] == "__county__"
            else el_q.get(r["did"], [])
        )
        for e in lst:
            for s in e["Secenekler"]:
                R = t.next_row()
                t.row(
                    [
                        r["name"],
                        KIND_LABEL.get(r["kind"], "İlçe"),
                        ELECTION_LABEL.get(e["Code"], e["Title"]),
                        e["Code"],
                        e["SandikSayisi"],
                        e["KayitliSecmen"],
                        e["KullanilanOy"],
                        e["GecerliOy"],
                        e["GecersizOy"],
                        s["Secenek"],
                        s["OySayisi"],
                        pct(f"K{R}", f"H{R}"),
                    ]
                )
    t.finish()

    # ================================================================ Hemşehri
    t = Table(
        wb.create_sheet("Hemşehri"),
        f"{place} — nüfusa kayıtlı il (ilk 10)",
        "Endeksa. Mahallede yaşayanların nüfus kütüğündeki il. Pay = kişi / mahalle nüfusu (Endeksa 2024), formül.",
        [
            Col("Sıra", "0", 6),
            Col("Kayıtlı il", "@", 16),
            Col("Kişi", FMT_INT, 9),
            Col("Mahalle nüfusu", FMT_INT, 10),
            Col("Pay", FMT_PCT, 8, True),
        ],
    )
    for r in rows + [{"name": "İLÇE", "did": "__county__", "q": county}]:
        f = (
            raw["fellows"]["county"]
            if r["did"] == "__county__"
            else fel_q.get(r["did"], {})
        )
        for k, x in enumerate((f or {}).get("FellowCountryman", []), 1):
            R = t.next_row()
            t.row(
                [
                    r["name"],
                    k,
                    x["CitizenCity"].title(),
                    x["CountOf"],
                    r["q"]["PopulationTotal"],
                    pct(f"D{R}", f"E{R}"),
                ]
            )
    t.finish()

    # ================================================================ TÜİK Nüfus
    t = Table(
        wb.create_sheet("TÜİK Nüfus"),
        f"{place} — TÜİK ADNKS mahalle nüfusu",
        "VeriAtlas ambarı (tuik_medas), yıl sonu nüfusu. Değişim, CAGR ve zirve yılı formül.",
        [Col(str(y), FMT_INT, 7.5, group="Toplam nüfus") for y in yrs]
        + [
            Col(f"Değişim {yrs[0]}→{yrs[-1]}", FMT_PCT, 10, True, "Özet"),
            Col("CAGR", "0.00%", 8, True, "Özet"),
            Col("Zirve yılı", "0", 8, group="Özet"),
        ],
    )
    for i, r in enumerate(rows):
        R = rr(i)
        a, b = f"B{R}", f"{L(1 + Y)}{R}"
        t.row(
            [r["name"]]
            + [tuik_get(r["tuik"], y) for y in yrs]
            + [
                growth(b, a),
                cagr(b, a, Y - 1),
                f'=IF(COUNT(B{R}:{L(1 + Y)}{R})>0,INDEX($B$4:${L(1 + Y)}$4,MATCH(MAX(B{R}:{L(1 + Y)}{R}),B{R}:{L(1 + Y)}{R},0))*1,"")',
            ]
        )
    add_subtotals(
        t,
        list(range(2, 2 + Y)),
        {},
        {
            2 + Y: f'=IF(AND(ISNUMBER(B{{r}}),B{{r}}>0),{L(1 + Y)}{{r}}/B{{r}}-1,"")',
            3
            + Y: f'=IF(AND(ISNUMBER(B{{r}}),B{{r}}>0),({L(1 + Y)}{{r}}/B{{r}})^(1/{Y - 1})-1,"")',
        },
        control=False,
    )
    # the "İLÇE" row has no Endeksa series; relabel as TÜİK sum of neighbourhoods
    t.ws.cell(row=R_COUNTY, column=1, value="İLÇE (mahalleler toplamı)")
    for c in range(2, 2 + Y):
        t.ws.cell(row=R_COUNTY, column=c, value=f"=SUM({L(c)}{R0}:{L(c)}{R0 + n - 1})")
    t.finish(R0 + n - 1)

    # ================================================================ TÜİK Yaş-Cinsiyet
    t = Table(
        wb.create_sheet("TÜİK Yaş-Cinsiyet"),
        f"{place} — TÜİK ADNKS 18+ / 0-17 × cinsiyet",
        "VeriAtlas ambarı. MEDAS mahalle düzeyinde yalnız bu iki yaş grubu var. Paylar formül.",
        [Col(str(y), FMT_INT, 7, group="18+") for y in yrs]
        + [Col(str(y), FMT_INT, 7, group="0-17") for y in yrs]
        + [Col(str(y), FMT_INT, 7, group="Erkek") for y in yrs]
        + [Col(str(y), FMT_INT, 7, group="Kadın") for y in yrs]
        + [
            Col(f"18+ {yrs[-1]}", FMT_PCT, 8, True, "Pay"),
            Col(f"0-17 {yrs[0]}", FMT_PCT, 8, True, "Pay"),
            Col(f"0-17 {yrs[-1]}", FMT_PCT, 8, True, "Pay"),
            Col(f"Kadın {yrs[-1]}", FMT_PCT, 8, True, "Pay"),
        ],
    )

    def tuik_age_shares(R: int) -> list:
        a18, c17, f17, kf = (
            f"{L(1 + Y)}{R}",
            f"{L(2 + Y)}{R}",
            f"{L(1 + 2 * Y)}{R}",
            f"{L(1 + 4 * Y)}{R}",
        )
        tl, tf = f"({a18}+{f17})", f"(B{R}+{c17})"
        return [
            f'=IF({tl}>0,{a18}/{tl},"")',
            f'=IF({tf}>0,{c17}/{tf},"")',
            f'=IF({tl}>0,{f17}/{tl},"")',
            f'=IF({tl}>0,{kf}/{tl},"")',
        ]

    for i, r in enumerate(rows):
        R = rr(i)
        vals = [r["name"]]
        for age, sex in (
            ("18+", "all"),
            ("0-17", "all"),
            ("all", "male"),
            ("all", "female"),
        ):
            vals += [tuik_get(r["tuik"], y, age, sex) for y in yrs]
        t.row(vals + tuik_age_shares(R))
    add_subtotals(
        t,
        list(range(2, 2 + 4 * Y)),
        {},
        {2 + 4 * Y + j: f for j, f in enumerate(tuik_age_shares(0))},
        control=False,
    )
    for R in (
        R_CENTRE,
        R_RURAL,
        R_COUNTY,
    ):  # fix {r}-less templates: rewrite shares with real row
        for j, f in enumerate(tuik_age_shares(R)):
            t.ws.cell(row=R, column=2 + 4 * Y + j, value=f)
    t.ws.cell(row=R_COUNTY, column=1, value="İLÇE (mahalleler toplamı)")
    for c in range(2, 2 + 4 * Y):
        t.ws.cell(row=R_COUNTY, column=c, value=f"=SUM({L(c)}{R0}:{L(c)}{R0 + n - 1})")
    t.finish(R0 + n - 1)

    # ================================================================ Kimlik
    t = Table(
        wb.create_sheet("Kimlik"),
        f"{place} — kimlik tablosu",
        "MEDAS kimliği: TÜİK / VeriAtlas anahtarı (il-ilçe-mahalle kodu); TÜİK serisi bu kodla bağlanır. Endeksa kimliği: Endeksa'nın kendi numarası. Tür: Merkez = ilçe merkezinin belediye mahalleleri, Kır = 2014 öncesi köy/belde.",
        [
            Col("MEDAS kimliği", "@", 18),
            Col("Endeksa kimliği", "0", 11),
            Col("Tür", "@", 8),
            Col("Endeksa verisi", "@", 10),
            Col("Sınır", "@", 7),
            Col("Belediye", "@", 18),
            Col("Yüzölçümü km²", "0.000", 11),
            Col("Nüfus 2024 TÜİK", FMT_INT, 11),
            Col("Nüfus 2024 Endeksa", FMT_INT, 11),
            Col("Fark", "0", 7),
        ],
    )
    for i, r in enumerate(rows):
        R = rr(i)
        t.row(
            [
                r["name"],
                r["area_id"],
                int(r["did"]),
                KIND_LABEL[r["kind"]],
                "yok" if r["ph"] else "var",
                "var" if r["did"] in raw["geo_ids"] else "yok",
                r["q"]["MunicipalityName"].title(),
                r["q"]["Area"],
                tuik_get(r["tuik"], 2024),
                r["q"]["PopulationTotal"],
                f'=IF(AND(ISNUMBER(I{R}),ISNUMBER(J{R})),J{R}-I{R},"")',
            ]
        )
    t.finish()

    # ================================================================ Özet (references the sheets above)
    t = Table(
        wb.create_sheet("Özet"),
        f"{place} — mahalle özeti",
        "Her sütun ilgili sayfadan formülle gelir; sayımlar kaynaktan. '(veri yok)' satırlarında Endeksa demografisi boş, TÜİK nüfus ve seçim geçerli.",
        [
            Col("Tür", "@", 8, group="Kimlik"),
            Col(f"Nüfus {yrs[-1]}", FMT_INT, 9, group="TÜİK"),
            Col(f"Değişim {yrs[0]}→{yrs[-1]}", FMT_PCT, 10, True, "TÜİK"),
            Col("Kadın payı", FMT_PCT, 8, False, "Yaş-cinsiyet"),
            Col("0-14 payı", FMT_PCT, 8, True, "Yaş-cinsiyet"),
            Col("65+ payı", FMT_PCT, 8, True, "Yaş-cinsiyet"),
            Col("Ort. yaş", FMT_DEC1, 8, True, "Yaş-cinsiyet"),
            Col("Yaşlanma endeksi", "0", 9, True, "Yaş-cinsiyet"),
            Col("Yüzölçümü km²", FMT_DEC2, 10, group="Alan"),
            Col("Yoğunluk kişi/km²", FMT_INT, 10, True, "Alan"),
            Col("Hane büyüklüğü", FMT_DEC2, 9, True, "Hane"),
            Col("Mülk sahibi", FMT_PCT, 9, group="Hane"),
            Col("Yüksek öğrenim payı", FMT_PCT, 9, True, "Eğitim"),
            Col("SES A+AB payı", FMT_PCT, 9, True, "SES"),
            Col("Hane geliri ₺/ay", FMT_INT, 10, True, "SES"),
            Col("Katılım", FMT_PCT, 8, True, "2024 İlçe Bld."),
            Col("1. parti", "@", 10, group="2024 İlçe Bld."),
            Col("1. parti payı", FMT_PCT, 8, True, "2024 İlçe Bld."),
        ],
    )

    def ozet_vals(R: int, kind: str, area_f, owner, income) -> list:
        return [
            kind,
            f"=Analiz!L{R}",
            f"=Analiz!M{R}",
            f"='Yaş Payları'!T{R}",
            f"='Yaş Payları'!B{R}",
            f"='Yaş Payları'!F{R}",
            f"=Analiz!B{R}",
            f"=Analiz!I{R}",
            area_f,
            f'=IF(AND(ISNUMBER(J{R}),J{R}>0,ISNUMBER(C{R})),C{R}/J{R},"")',
            f"=Analiz!J{R}",
            owner,
            f"=Eğitim!O{R}",
            f"=SES!H{R}",
            income,
            f"='Seçim 2024'!G{R}",
            f"='Seçim 2024'!{L(TOPC)}{R}",
            f"='Seçim 2024'!{L(TOPS)}{R}",
        ]

    for i, r in enumerate(rows):
        R = rr(i)
        t.row(
            [r["name"] + (" (veri yok)" if r["ph"] else "")]
            + ozet_vals(
                R,
                KIND_LABEL[r["kind"]],
                r["q"]["Area"],
                None if r["ph"] else r["q"]["OwnerShare"] / 100,
                None if r["ph"] else r["q"]["HouseIncomeTotal"],
            )
        )
    for R, label, idx, style in (
        (R_CENTRE, "MERKEZ", centre_rows, "subtotal"),
        (R_RURAL, "KIR", rural_rows, "subtotal"),
    ):
        t.row([label] + ozet_vals(R, "", sum_rows("J", idx), None, None), style=style)
    t.row(
        ["İLÇE (Endeksa)"]
        + ozet_vals(
            R_COUNTY,
            "",
            county["Area"],
            county["OwnerShare"] / 100,
            county["HouseIncomeTotal"],
        ),
        style="total",
    )
    t.ws.cell(row=R_COUNTY, column=3, value=county["PopulationTotal"])
    t.ws.cell(row=R_COUNTY, column=4, value=None)
    t.finish(R0 + n - 1)

    # ================================================================ Ham
    keys = list(county)
    t = Table(
        wb.create_sheet("Ham"),
        f"{place} — Endeksa demografi yanıtı, tüm alanlar",
        "Alan adları kaynaktaki gibi; ilk satır ilçe. Diğer sayfalarda kullanılmayan alanlar da burada.",
        [Col(k, "General", 12) for k in keys[1:]],
        first_col=Col(keys[0], "General", 8),
    )
    t.row([county.get(k) for k in keys], style="total")
    for r in rows:
        t.row([r["q"].get(k) for k in keys])
    t.finish()

    # ================================================================ Notlar
    ws = wb.create_sheet("Notlar")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Notlar"
    ws["B2"].font = Font(name=FONT, bold=True, size=14, color=NAVY)
    notes = [
        (
            "Kaynaklar",
            f"Endeksa (endeksa.com) demografi, seçim, hemşehri, sınır uç noktaları; döküm {fetched}. TÜİK ADNKS mahalle serisi VeriAtlas ambarından (public/population-neighbourhood.csv.gz, kaynak tuik_medas).",
        ),
        (
            "Yıl",
            "Endeksa tek kesittir; nüfusu TÜİK ADNKS 2024 ile birebir. Seçimler kendi tarihlerinde. Emlak satış serileri yıllık (2024 kısmi).",
        ),
        (
            "Veri yok",
            "Küçük eski köylerde Endeksa mahalle düzeyi demografi üretmiyor (hane = 0 şablonu). Bu mahalleler '(veri yok)' etiketli; TÜİK nüfus, seçim ve hemşehri onlarda da dolu. Merkez/Kır toplamları yalnız veri olan mahalleleri içerir; Kır toplamı bu yüzden TÜİK kır nüfusundan küçüktür.",
        ),
        (
            "Tahmin",
            "SES, gelir, harcama, tasarruf, mülk/kiracı Endeksa modelidir; TÜİK sayımı değildir. Yaş, cinsiyet, eğitim, medeni hal TÜİK ADNKS kökenli görünüyor (toplamlar tutuyor).",
        ),
        (
            "Ortalama yaş",
            "5'lik grup orta noktaları (2, 7, …, 62) ve 65+ için 74 ile ağırlıklı ortalama; 65+ grubunun gerçek ortalaması mahalleye göre değişir, ±1 yıl belirsizlik. Medyan yaş grup içi doğrusal ara değerle Python'da hesaplandı (formül değil).",
        ),
        (
            "Tür",
            "'Merkez' = Endeksa kimliği < 100000 (ilçe merkezinin belediye mahalleleri); 'Kır' = 2014 öncesi köy ve belde. Belde ayrımı bu sürümde yapılmadı (İznik: Boyalıca, Elbeyli eski belde).",
        ),
        (
            "Formüller",
            "Pay, yoğunluk, ortalama yaş, katılım, 1. parti gibi sütunlar formüldür; bir sayım düzeltilirse güncellenir. Dosya açılınca hesaplanır. Aynı mahalle her sayfada aynı satırdadır; Özet diğer sayfalara bu satır numarasıyla bağlanır.",
        ),
        (
            "Seçim",
            "Küçük partiler kaynakta 'Diğer' altında; tam döküm YSK'dan. Kaynaktaki 'Oran' alanı hatalı olduğundan kullanılmadı; paylar geçerli oya bölünerek yeniden hesaplandı.",
        ),
        (
            "Kimlikler",
            "MEDAS kimliği (TR-il-ilçe-kod) VeriAtlas anahtarı; TÜİK serisiyle bu kodla birleşir. Endeksa kimliği yalnız Endeksa'ya geri dönmek için.",
        ),
        ("Lisans", "Endeksa verisi araştırma amaçlı kopyadır; yayım kararı verilmedi."),
    ]
    for i, (k, v) in enumerate(notes, 4):
        ws.cell(row=i, column=2, value=k).font = Font(
            name=FONT, bold=True, size=10, color=NAVY
        )
        c = ws.cell(row=i, column=3, value=v)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 48
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 120

    desired = [
        "Kapak",
        "Özet",
        "Analiz",
        "Kimlik",
        "Yaş Toplam",
        "Yaş Erkek",
        "Yaş Kadın",
        "Yaş Payları",
        "Eğitim",
        "Medeni Hal",
        "SES",
        "Gelir-Harcama",
        "Konut",
        "Emlak Satış",
        "Seçim 2024",
        "Katılım",
        "Seçimler",
        "Hemşehri",
        "TÜİK Nüfus",
        "TÜİK Yaş-Cinsiyet",
        "Ham",
        "Notlar",
    ]
    wb._sheets = [wb[nm] for nm in desired]
    tab = {
        "Kapak": NAVY,
        "Özet": NAVY,
        "Analiz": NAVY,
        "Kimlik": "7F7F7F",
        "Ham": "7F7F7F",
        "Notlar": "7F7F7F",
    }
    for sh in wb.worksheets:
        sh.sheet_properties.tabColor = tab.get(
            sh.title,
            BLUE
            if sh.title.startswith(("Yaş", "Eğitim", "Medeni"))
            else "548235"
            if sh.title.startswith(("Seçim", "Katılım"))
            else "BF8F00",
        )
    wb.active = 0
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--district", required=True, help="VeriAtlas district id, e.g. TR-16-006"
    )
    ap.add_argument(
        "--dump", type=Path, help="single endeksaFetch.download() JSON to unpack first"
    )
    ap.add_argument("--raw", type=Path, default=Path("C:/veri/raw/endeksa"))
    ap.add_argument("--out", type=Path)
    a = ap.parse_args()
    raw_dir = a.raw / a.district
    if a.dump:
        unpack_dump(a.dump, raw_dir)
    raw = load_raw(raw_dir)
    areas, series = load_tuik(a.district)
    out = a.out or (a.raw / f"{a.district}-endeksa.xlsx")
    build(raw, areas, series, out)
    print(out)


if __name__ == "__main__":
    main()
